from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .lda_rules import calculate_summary, enrich_records_for_analysis

BASE_DIR = Path(__file__).resolve().parents[1]
EXPORT_DIR = BASE_DIR / "exports"


def export_interview_xlsx(interview: Dict[str, Any], records: List[Dict[str, Any]], standard_minutes: float) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    cargo_safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(interview.get("cargo") or "cargo"))[:40]
    out_path = EXPORT_DIR / f"LDA_{cargo_safe}_entrevista_{interview.get('id')}.xlsx"

    enriched = enrich_records_for_analysis(records, standard_minutes)
    summary = calculate_summary(enriched, standard_minutes)
    df = pd.DataFrame(enriched)
    desired_cols = [
        "cod", "funcion_origen", "proceso", "actividad", "tipo_actividad", "pregunta_validacion",
        "evidencia_esperada", "herramienta_sistema", "entregable", "indicador", "criticidad", "automatizable", "riesgo", "dependencia",
        "semana_1", "semana_2", "semana_3", "semana_4", "diario", "semanal", "quincenal", "mensual", "anual",
        "tiempo_x_unidad_min", "cumplimiento", "evidencia", "comentarios_usuario", "comentarios_proyecto", "opt",
        "vol_mes", "val", "min_mes", "hrs_mes", "min_dia", "fte_pct", "cumplimiento_score", "riesgo_score",
        "potencial_ahorro_min_dia", "hallazgo", "errores_validacion",
    ]
    for col in desired_cols:
        if col not in df.columns:
            df[col] = ""
    df = df[desired_cols]

    summary_rows = [
        ["Empresa", interview.get("company_name", "")],
        ["Cargo", interview.get("cargo", "")],
        ["Área", interview.get("area", "")],
        ["Consultor", interview.get("consultant_name", "")],
        ["Entrevistado", interview.get("entrevistado", "")],
        ["Fecha", interview.get("fecha", "")],
        ["Tiempo estándar diario", summary["tiempo_estandar"]],
        ["Minutos día reportados", summary["min_dia_total"]],
        ["Brecha min/día", summary["brecha_min_dia"]],
        ["Utilización %", summary["utilizacion_pct"]],
        ["Cumplimiento ponderado %", summary["cumplimiento_ponderado_pct"]],
        ["Core min/día", summary["min_core_dia"]],
        ["No core min/día", summary["min_no_core_dia"]],
        ["Actividades válidas", summary["actividades_validas"]],
        ["Actividades con error", summary["actividades_invalidas"]],
        ["Oportunidades", summary["oportunidades"]],
        ["Ahorro potencial min/día", summary["potencial_ahorro_min_dia"]],
        ["Riesgo promedio", summary["riesgo_promedio"]],
        ["Conclusión", interview.get("conclusion", "")],
    ]

    process_df = pd.DataFrame()
    if not df.empty:
        valid = df[df["val"] == 1]
        if not valid.empty:
            process_df = valid.groupby("proceso", dropna=False).agg(
                min_dia=("min_dia", "sum"),
                hrs_mes=("hrs_mes", "sum"),
                actividades=("actividad", "count"),
                ahorro_potencial=("potencial_ahorro_min_dia", "sum"),
            ).reset_index().sort_values("min_dia", ascending=False)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows, columns=["Indicador", "Valor"]).to_excel(writer, sheet_name="Resumen", index=False)
        df.to_excel(writer, sheet_name="LDA_enriquecida", index=False)
        process_df.to_excel(writer, sheet_name="Procesos", index=False)

    _style_export(out_path)
    return out_path


def _style_export(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col[:80])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)
    wb.save(path)
